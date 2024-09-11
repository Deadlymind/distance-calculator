import openpyxl

wb = openpyxl.load_workbook("suppliers.xlsx")
ws = wb['Sheet1']


product_name = ws.cell(row=3, column=4).value
product_unit = ws.cell(row=4, column=4).value
product_margin = ws.cell(row=5, column=4).value

data_start = 11
data_end = 56

suppliers_data = []

for row in range(data_start, data_end):
    data = {
        "date": ws.cell(row=row, column=1).value,
        "identifier": ws.cell(row=row, column=2).value,  #EMPTY
        "company_name": ws.cell(row=row, column=3).value,
        "contact_name": ws.cell(row=row, column=4).value,
        "landline": ws.cell(row=row, column=5).value,
        "mobile": ws.cell(row=row, column=6).value,
        "email": ws.cell(row=row, column=7).value,
        "country": ws.cell(row=row, column=8).value,
        "department": ws.cell(row=row, column=9).value,
        "current_price": ws.cell(row=row, column=10).value,
        "price_type": ws.cell(row=row, column=11).value,
        "departure_address": ws.cell(row=row, column=12).value,
        "arrival_address": ws.cell(row=row, column=13).value,
        "number_of_distance": ws.cell(row=row, column=14).value,  #EMPTY
        "transport_price": ws.cell(row=row, column=15).value,  #EMPTY
        "weight_of_full_truck": ws.cell(row=row, column=16).value,
        "format": ws.cell(row=row, column=17).value,
        "comments": ws.cell(row=row, column=18).value,
    }
    suppliers_data.append(data)

print()