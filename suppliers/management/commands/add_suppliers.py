import openpyxl
from geopy.geocoders import Nominatim
from datetime import datetime

from django.core.management.base import BaseCommand

from countries.models import Regions
from products.models import Product, Units
from suppliers.models import Suppliers, SupplierPrices
from accounts.models import User
from suppliers.utils import calculate_price_exwb_from_exws, calculate_price_exws_from_daps, \
    calculate_road_distance_price
from offers.utils import generate_offers
from countries.utils import get_address_coordinates_from_mapbox, \
    closest_location


data_start = 11
data_end = 313
filename = "suppliers-v2.xlsx"
sheet_name = 'Feuille 1'

def parse_date(item):
    return str(item['date'])


class Command(BaseCommand):

    def scrap_data(self):
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name]

        self.product_name = ws.cell(row=3, column=4).value
        self.product_unit = ws.cell(row=4, column=4).value
        self.product_margin = ws.cell(row=5, column=4).value

        self.suppliers_data = []

        for row in range(data_start, data_end):
            data = {}
            data["date"]= ws.cell(row=row, column=1).value
            if ws.cell(row=row, column=2).value:
                data["identifier"]= ws.cell(row=row, column=2).value  #EMPTY
            else:
                data["identifier"] = ""
            data['row'] = row
            data["company_name"]= ws.cell(row=row, column=3).value
            data["contact_name"]= ws.cell(row=row, column=4).value
            data["landline"]= ws.cell(row=row, column=5).value
            data["mobile"]= ws.cell(row=row, column=6).value
            data["email"]= ws.cell(row=row, column=7).value
            data["country"]= ws.cell(row=row, column=8).value
            data["department"]= ws.cell(row=row, column=9).value
            data["current_price"]= ws.cell(row=row, column=10).value
            data["price_type"]= ws.cell(row=row, column=11).value
            data["departure_address"]= ws.cell(row=row, column=12).value
            data["arrival_address"]= ws.cell(row=row, column=13).value
            data["number_of_distance"]= ws.cell(row=row, column=14).value  #EMPTY
            data["transport_price"]= ws.cell(row=row, column=15).value #EMPTY
            if ws.cell(row=row, column=16).value:
                data["weight_of_full_truck"] = ws.cell(row=row, column=16).value
            else:
                data["weight_of_full_truck"] = 25
            data["format"]= ws.cell(row=row, column=17).value
            data["comments"]= ws.cell(row=row, column=18).value
            self.suppliers_data.append(data)
        
        sorted_data = sorted(self.suppliers_data, key=parse_date)
        self.suppliers_data = sorted_data

    def get_unit(self):
        unit = Units.objects.get_or_create(name=self.product_unit)[0]
        print(f"UNIT: {unit}")
        return unit

    def get_product(self):
        product = Product.objects.filter(name=self.product_name.lower())
        if product:
            product = product.first()
            print(f"PRODUCT: {product}")
            product.default_margin = self.product_margin
            product.default_unit = self.get_unit()
            product.save()
        else:
            product = Product.objects.create(
                name=self.product_name.lower(),
                default_margin=self.product_margin,
                default_unit=self.get_unit(),
            )
        return product

    def get_region(self, lat, lng):
        return closest_location(lat, lng, self.regions)

    def get_address_coordinates(self, address):
        location = get_address_coordinates_from_mapbox(address)
        if location:
            return {
                "latitude": location['lat'],
                "longitude": location['lng'],
            }
        else:
            raise Exception("ADDRESS NOT FOUND")

    def get_arrival_address_coordinates(self, address):
        if "None" in address:
            raise Exception(f"ARRIVAL ADDRESS CANNOT BE NULL ON DAPS")
        geolocator = Nominatim(user_agent="Map")
        location = geolocator.geocode(address)
        if location:
            return {
                "latitude": location.latitude,
                "longitude": location.longitude,
            }
        else:
            print("OSM ADDRESS NOT FOUND, NOW USING MAPBOX TO FIND ADDRESS")
            location_mapbox = get_address_coordinates_from_mapbox(address)
            return {
                "latitude": location_mapbox['lat'],
                "longitude": location_mapbox['lng'],
            }
            # raise Exception(f"ADDRESS NOT FOUND {address}")


    def handle(self, *args, **options):
        self.regions = Regions.objects.all()
        self.suppliers_data = []
        self.scrap_data()
        product = self.get_product()
        print(product)
        added_by = User.objects.first()

        for supplier_data in self.suppliers_data:
            try:
                if Suppliers.objects.filter(company_name=supplier_data['company_name']).exists():
                    supplier = Suppliers.objects.get(company_name=supplier_data['company_name'])
                    print(f"DUPLICATE SUPPLIER: {supplier}")
                else:
                    supplier = Suppliers(
                        product=product,
                        added_by=added_by,
                        company_name=supplier_data['company_name'],
                        contact_name=supplier_data['contact_name'],
                        landline=supplier_data['landline'],
                        mobile=supplier_data['mobile'],
                        email=supplier_data['email'],

                        identifier=supplier_data['identifier'],
                        format=supplier_data['format'],
                        comments=supplier_data['comments'],

                        weight_of_full_truck=float(supplier_data.get('weight_of_full_truck')),
                        margin=product.default_margin,
                    )
                    print(f"SUPPLIER CREATED: {supplier.company_name}, {supplier_data['date']}")

                    address = self.get_address_coordinates(
                        supplier_data['departure_address']
                    )
                    supplier.departure_address_lat = address['latitude']
                    supplier.departure_address_lng = address['longitude']

                    supplier.region = self.get_region(
                        supplier.departure_address_lat,
                        supplier.departure_address_lng
                    )

                if supplier_data['price_type'] == "EXWS":
                    exws = supplier_data['current_price']
                    exwb = calculate_price_exwb_from_exws(exws, product.default_margin)

                elif supplier_data['price_type'] == "DAPS":
                    arrival_address = self.get_arrival_address_coordinates(
                        f"{supplier_data['arrival_address']}, {supplier_data['department']}, {supplier_data['country']}"
                    )
                    distance_price = calculate_road_distance_price(
                        [supplier.departure_address_lat, supplier.departure_address_lng],
                        [arrival_address['latitude'], arrival_address['longitude']],
                        supplier.region
                    )
                    exws = calculate_price_exws_from_daps(
                        supplier_data['current_price'], distance_price, supplier_data['weight_of_full_truck']
                    )
                    exwb = calculate_price_exwb_from_exws(exws, product.default_margin)

                supplier.price_exws = exws
                supplier.price_exwb = exwb

                supplier.save()

                prices = SupplierPrices(
                    suppliers=supplier,
                    price_exwb=supplier.price_exwb,
                    price_exws=supplier.price_exws,
                    added_by=added_by,
                    added_on=datetime.strptime(supplier_data['date'], "%d/%m/%Y") if isinstance(supplier_data['date'], str) else supplier_data['date']
                )
                prices.save()
            except Exception as error:
                print(f"ERROR OCCURED FOR ROW # {supplier_data['row']}: {error}")
        generate_offers()